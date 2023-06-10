from hijri_converter import convert
from openpyxl import load_workbook, Workbook
from datetime import datetime

# Function to convert Hijri date to Gregorian date
def convert_hijri_to_gregorian(hijri_date):
    try:
        hijri_date_parts = hijri_date.split("/")
        hijri_year = int(hijri_date_parts[2])
        hijri_month = int(hijri_date_parts[1])
        hijri_day = int(hijri_date_parts[0])
        gregorian_date = str(convert.Hijri(hijri_year, hijri_month, hijri_day).to_gregorian()).split("-")
        gregorian_date_new = gregorian_date[2]+"/"+gregorian_date[1]+"/"+gregorian_date[0]
        return gregorian_date_new
    except Exception as e:
        print(f"Error converting Hijri date {hijri_date}: {str(e)}")
        return None


def time_diff(end, start):
    try:
        end_date = datetime.strptime(end, "%d/%m/%Y")
        start_date = datetime.strptime(start, "%d/%m/%Y")
        diff = end_date - start_date
        return diff.days
    except Exception as e:
        print(f"Error calculating time difference: {str(e)}")
        return None


# Load the Excel file
input_file_path = "Book1.xlsx"
output_file_path = "output_dates.xlsx"

# # Open the input file and get the active sheet
input_wb = load_workbook(input_file_path)
input_sheet = input_wb.active

# Create a new workbook for the output
output_wb = Workbook()
output_sheet = output_wb.active

# Write headers in the output file
output_sheet["A1"] = "Hijri Start Date"
output_sheet["B1"] = "Gregorian Start Date"
output_sheet["C1"] = "Hijri End Date"
output_sheet["D1"] = "Gregorian End Date"
output_sheet["F1"] = "Duration"

# Create a new workbook and get the active sheet
wb = Workbook()
sheet = wb.active

for i, hijri_date in enumerate(input_sheet.iter_rows(values_only=True), start=2):
    # hijri_date = row[0]
    _time_diff = None
    gregorian_date_start = None
    gregorian_date_end = None

    if hijri_date[0] is not None:
        print(hijri_date[0])
        gregorian_date_start = convert_hijri_to_gregorian(hijri_date[0])
        output_sheet.cell(row=i, column=1, value=hijri_date[0])
        output_sheet.cell(row=i, column=2, value=gregorian_date_start)

    if hijri_date[2] is not None:
        print(hijri_date[2])
        gregorian_date_end = convert_hijri_to_gregorian(hijri_date[2])
        output_sheet.cell(row=i, column=3, value=hijri_date[2])
        output_sheet.cell(row=i, column=4, value=gregorian_date_end)

    if gregorian_date_end and gregorian_date_start is not None: 
        _time_diff = time_diff(gregorian_date_end, gregorian_date_start)
    if _time_diff is not None:
        output_sheet.cell(row=i, column=6, value=_time_diff)

# Save the output file
output_wb.save(output_file_path)
