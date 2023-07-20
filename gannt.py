from openpyxl import load_workbook
from datetime import datetime

def convert_to_iso_format(date_str):
    # Try parsing the date string with different formats
    dt_object = None
    formats_to_try = ['%d/%m/%Y',
                    #    '%Y-%m-%d',
                    #    '%d-%m-%Y',
                    '%Y-%d-%m %H:%M:%S']
    for date_format in formats_to_try:
        try:
            dt_object = datetime.strptime(str(date_str), date_format)
            # date_obj = datetime.strptime(str(date_str), '%d/%m/%Y')
            break
        except ValueError:
            continue

    if dt_object is None:
        # raise ValueError("Invalid date format")
        return None

    # Set the time to 5 AM
    new_date_str = dt_object.strftime('%m/%d/%Y')
    # dt_object = dt_object.replace(hour=5, minute=0, second=0, microsecond=0)

    # Format the datetime object into ISO 8601 format
    # iso_date_str = dt_object.isoformat()
    return new_date_str

def convert_date_format(input_file):
    # Open the Excel file using openpyxl
    wb = load_workbook(input_file)
    ws = wb.active

    # Iterate through the date column and convert the dates
    # for row in ws.iter_rows(min_row=8, values_only=True):
    #     date_str = row[4]  # Assuming the date column is in the second (B) column
    #     if date_str is not None:
            
    #         # date_obj = datetime.strptime(str(date_str), '%d/%m/%Y')
    #         new_date_str = convert_to_iso_format(date_str)
    #         print(new_date_str)
    #         cell = ws.cell(row=row[0], column=5)  # Assuming the date column is in the second (B) column
    #         cell.value = new_date_str
            # row[4] = new_date_st

    # Assuming the date column is in the second (B) column
    date_column = ws['E']

    # Iterate through the date column and convert the dates
    for cell in date_column[1:]:
        date_str = cell.value
        if date_str is not None:
            new_date_str = convert_to_iso_format(date_str)
            if new_date_str is not None:
                print(new_date_str)
                cell.value = new_date_str
            else:
                print('error')
        else:
            print('empty')

    # Save the changes back to the same Excel file
    wb.save(input_file)

if __name__ == "__main__":
    # Replace 'input.xlsx' with the name of your input Excel file.
    convert_date_format("test.xlsx")
