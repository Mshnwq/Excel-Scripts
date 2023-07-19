import os
import requests
import time
import json
from datetime import datetime
from openpyxl import load_workbook

# Load the ACCESS_TOKEN from the .env file
from dotenv import load_dotenv

load_dotenv()
ACCESS_TOKEN = os.getenv("ACCESS_TOKEN")

# Function to make the API request
def add_event_to_calendar(event_body):
    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }

    endpoint = "https://graph.microsoft.com/v1.0/me/events"

    try:
        response = requests.post(endpoint, headers=headers, json=event_body)

        if response.status_code == 200 or response.status_code == 201:
            print("Event added successfully.")
        else:
            print(f"Failed to add event. Status code: {response.status_code}, Error: {response.text}")
    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")

def convert_to_iso_format(date_str):
    # Try parsing the date string with different formats
    dt_object = None
    formats_to_try = ['%d/%m/%Y',
                    #    '%Y-%m-%d',
                    #    '%d-%m-%Y',
                    '%Y-%d-%m %H:%M:%S']
    for date_format in formats_to_try:
        try:
            dt_object = datetime.strptime(str(date_str), date_format)
            break
        except ValueError:
            continue

    if dt_object is None:
        raise ValueError("Invalid date format")
        # return None

    # Set the time to 5 AM
    dt_object = dt_object.replace(hour=5, minute=0, second=0, microsecond=0)

    # Format the datetime object into ISO 8601 format
    iso_date_str = dt_object.isoformat()
    return iso_date_str

if __name__ == "__main__":
    # Load the Excel file
    wb = load_workbook("dates-3.xlsx")
    sheet = wb.active
    events_list = []

    # Add the events to the calendar
    for row in sheet.iter_rows(min_row=7, values_only=True):  # Assuming until row 6 contains headers
        title, start_date, end_date = row[1], row[5], row[8]
        print(f"title: {title}")
        print(f"start: {start_date}")
        print(f"end: {end_date}")
        if title is not None and start_date is not None and end_date is not None:

            start_date_iso = convert_to_iso_format(start_date)
            end_date_iso = convert_to_iso_format(end_date)

            if start_date_iso is not None and end_date_iso is not None:
                event_body = {
                    "subject": title,
                    "start": {
                        "dateTime": start_date_iso,
                        "timeZone": "UTC"
                    },
                    "end": {
                        "dateTime": end_date_iso,
                        "timeZone": "UTC"
                    }
                }
                events_list.append(event_body)
                # break
            else:
                print("\nfailed iso\n")

    # Convert the events_list to a JSON string with indentation for readability
    events_json_str = json.dumps(events_list, indent=4)

    # Print the JSON in a nice format
    print(events_json_str)
    for event in events_list:
        add_event_to_calendar(event)
        time.sleep(0.2)
